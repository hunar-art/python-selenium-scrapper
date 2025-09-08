import pandas as pd
import requests
import time
import os
from dotenv import load_dotenv
from mongodb import MongoClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

load_dotenv()
GoogleAPI = os.getenv("GoogleAPI")
outputData = "mongodbDataVsGMV03.xlsx"

uri = "mongodb+srv://YOUR CONNECTION STRING HERE"
client = MongoClient(uri)
db = client["iHealth_Dev"]
collection = db["ProviderLookupCollection"]

mongo_data_df = pd.DataFrame(list(collection.find()))

us_states = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA",
    "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "Florida": "FL", "Georgia": "GA",
    "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH",
    "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY", "North Carolina": "NC",
    "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA",
    "Rhode Island": "RI", "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN",
    "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY", "District of Columbia": "DC"
}
none_data = ["nan", "none", "", "null"]

def normalize_phone(phone):
    if not phone:
        return ""
    digits = re.sub(r'\D', '', str(phone))
    if len(digits) == 10:        
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    return phone

def extracted_root_domain(url):
    if not url or str(url).lower() in none_data:
        return ""
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    parts = url.split('.')
    if len(parts) >= 2:
        return '.'.join(parts[-2:])
    return url

def verify_hospital(clinic_name):
    search_url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    details_url = "https://maps.googleapis.com/maps/api/place/details/json"

    search_params = {"query": clinic_name, "key": GoogleAPI}
    search_res = requests.get(search_url, params=search_params).json()
    print("Search response:", search_res)  # Debug line
    if search_res["status"] != "OK" or not search_res["results"]:
        return None

    place_id = search_res["results"][0]["place_id"]
    place_types = search_res["results"][0].get("types", [])

    details_params = {
        "place_id": place_id,
        "fields": "name,formatted_address,formatted_phone_number,types,website",
        "key": GoogleAPI,
    }
    details_res = requests.get(details_url, params=details_params).json()
    print("Details response: ", details_res)  # Debug line
    if details_res["status"] != "OK":
        return None

    result = details_res["result"]
    result["types"] = result.get("types", place_types) 
    print("Verified result:", result)  # Debug line
    return result

# ----------------------
def value_compration(new,old):
    if str(new).strip().lower() == str(old).strip().lower():
        return "No Changes"
    if str(new) is None or str(new).strip().lower() in none_data and str(old):
        return "Missed"
    elif old is None or str(old).strip().lower() in none_data:
        return "Addition"
    elif (new and old) and str(new).lower().strip != str(old).lower().strip():
        return "Change"


for idx, row in mongo_data_df.iterrows():
    hospital = str(row.get("hospital", "")).strip()
    physician = str(row.get("physician_1", "")).strip()

    query = None
    if hospital.lower() not in none_data and physician.lower() not in none_data:
        query = f"{hospital} {physician}"
    elif hospital.lower() in none_data and physician.lower() not in none_data:
        query = physician
    elif physician.lower() in none_data and hospital.lower() not in none_data:
        query = f"{hospital} hospital" 

    if not query:
        print("Skipping row with no valid query")
        continue
    
    # print(f"Verifying query: {query}")    
    verified = verify_hospital(query)
    print("Verified query:", verified)
    if verified:
        mongo_data_df.at[idx, "physicians_1_NEW"] = verified.get("name", "")
        mongo_data_df.at[idx, "physicians_1_difference"] = value_compration(verified.get("name", ""),row["physician_1"])


        formatted_phone = normalize_phone(verified.get("formatted_phone_number", ""))
        mongo_data_df.at[idx, "phone_NEW"] = formatted_phone
        mongo_data_df.at[idx, "phone_difference"] = value_compration(verified.get("formatted_phone_number", ""),normalize_phone(row["phone_1"]))

        # Address parsing
        addr = verified.get("formatted_address", "")
        print(f"Address: {addr}")
        parts = [a.strip() for a in addr.split(",")]

        street, city, state, zip_code = "", "", "", ""

        state_codes = {v: k for k, v in us_states.items()}


        if len(parts) >= 3:
            state_zip = parts[-2].split()
            if len(state_zip) >= 2:
                state, zip_code = state_zip[0], state_zip[1]

            city = parts[-3]

            street = ", ".join(parts[:-3])
        print('--------------------------------------------------')

        print(f"Parsed: street='{street}', city='{city}', state='{state}', zip='{zip_code}'")
        mongo_data_df.at[idx, "street_address_NEW"] = street
        mongo_data_df.at[idx, "street_address_difference"] = value_compration(street,row["street_address"])
        mongo_data_df.at[idx, "provider_city_NEW"] = city
        mongo_data_df.at[idx, "provider_city_difference"] = value_compration(city,row["provider_city"])
        mongo_data_df.at[idx, "provider_state_NEW"] = state
        
        old_state = str(row["provider_state"]).strip()
        new_state = state.strip()

        if old_state in us_states:
            old_state = us_states[old_state]

        old_state = old_state.upper()
        new_state = new_state.upper()

        mongo_data_df.at[idx, "provider_state_difference"] = value_compration(new_state, old_state)

        mongo_data_df.at[idx, "provider_zip_code_NEW"] = zip_code
        mongo_data_df.at[idx, "provider_zip_code_difference"] = value_compration(zip_code,row["provider_zip_code"])

        mongo_data_df.at[idx, "website_NEW"] = verified.get("website", "")

        mongo_data_df.at[idx, "website_difference"] = value_compration(extracted_root_domain(verified.get("website", "")),extracted_root_domain(row["website"]))

        specialty_list = verified.get("types", [])
        specialty_str = ", ".join(specialty_list)
        mongo_data_df.at[idx, "specialty_NEW"] = specialty_str
        mongo_data_df.at[idx, "specialty_difference"] = value_compration(specialty_str,row["specialty"])

        time.sleep(1)
        # print(f"Processed, Address:  {mongo_data_df['street_address_NEW'].iloc[idx]}, city name:  {mongo_data_df['provider_city_NEW'].iloc[idx]}, State: {mongo_data_df['provider_state_NEW'].iloc[idx]}, Zip code:  {mongo_data_df['provider_zip_code_NEW'].iloc[idx]}")
    else:
        print(f"Could not verify {physician}")


mongo_data_df.to_excel(outputData, index=False)
print("Verification completed")

print("-----------------------")
print(f"Applying highlights to the excel file")
workbook = load_workbook(outputData)
worksheet = workbook.active
fills = {
    "Change": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),   # Yellow
    "No Changes": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light Blue
    "Addition": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light Green
    "Missed": PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid"),    # Tomato Red
}
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        header = worksheet.cell(row=1, column=cell.column).value
        if header and "_difference" in str(header):
            value = str(cell.value).strip()
            if value in fills:
                cell.fill = fills[value]

workbook.save(outputData)

print("applied conditional formatting")
